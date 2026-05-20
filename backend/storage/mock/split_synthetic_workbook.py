from pathlib import Path

from openpyxl import Workbook, load_workbook


SOURCE = Path("/app/storage/mock/module_g_synthetic_smoke.xlsx")
TARGET_DIR = Path("/app/storage/mock/split")
MAPPING = {
    "Wells": "wells.xlsx",
    "GTM": "gtm.xlsx",
    "Infrastructure": "infrastructure.xlsx",
    "KRS": "external_krs_schedule.xlsx",
}


def copy_sheet(source_path: Path, sheet_name: str, output_path: Path) -> None:
    source_wb = load_workbook(source_path)
    source_ws = source_wb[sheet_name]

    target_wb = Workbook()
    target_ws = target_wb.active
    target_ws.title = sheet_name

    for row in source_ws.iter_rows():
        for cell in row:
            target_ws[cell.coordinate] = cell.value

    target_wb.save(output_path)
    source_wb.close()
    target_wb.close()


def main() -> None:
    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    for sheet_name, filename in MAPPING.items():
        copy_sheet(SOURCE, sheet_name, TARGET_DIR / filename)
        print(f"{sheet_name} -> {filename}")


if __name__ == "__main__":
    main()
